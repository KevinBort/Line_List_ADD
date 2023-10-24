import os, re, openpyxl
from openpyxl.styles import Alignment
import pandas as pd

current_path=os.getcwd()
input_file_name='para_line_list.xlsx'
input_file_path=os.path.join(current_path,input_file_name)

# TODO ESTO ES COMPLETAMENTE AL PEDO PERO LO HAGO PARA PRACTICAR.... PODRIA DIRECTAMENTE ARMAR EL DICCIONARIO YO Y LISTO, HASTA QUIZAS SEA MEJOR PARA EVITAR PROBLEMAS CON EL EXCEL.. VER PARA CAMBIAR EN ALGUN MOMENTO.
# 
# ------------------------------------------------------------------------------------------------------------------ 
wbook=openpyxl.load_workbook(input_file_path)                                                                      
wsheet=wbook.active

medidas=dict()
medidas.update({'OD8mm':'8','OD12mm':'12'})



# # Armo un diccionario con los valores de las medidas en pulgadas y su DN en mm correspondiente.
for row in wsheet.iter_rows(min_row=2,max_row=16,min_col=1,max_col=2,values_only=True):
    medidas.update({row[1]:row[0]})

print(medidas)
wbook.save('para_line_list.xlsx')
# ------------------------------------------------------------------------------------------------------------------ 
# # Cargo el template para sobreescribir el line list. 
line_list_template=openpyxl.load_workbook(os.path.join(current_path,'line_list_template.xlsx'))
wsheet_ll=line_list_template.active
# print (wsheet_ll.cell(row=11,column=4).value)
row_start=11

columna_NPS=wsheet_ll['D']
columna_Pdis=wsheet_ll['T']
columna_Fluido=wsheet_ll['E']
columna_fase=wsheet_ll['Q']
for row_index in range(row_start, wsheet_ll.max_row + 1): # Para todas las celdas escribir tabla, estable, y grupo.
    cell_NPS = columna_NPS[row_index - 1]
    cell_Pdis = columna_Pdis[row_index - 1]
    cell_Fluido = columna_Fluido[row_index - 1]
    cell_fase = columna_fase[row_index - 1]
    wsheet_ll.cell(row=row_index,column=35,value='NO')   # Ver si no hay que poner algún condicional para estas líneas. 
    wsheet_ll.cell(row=row_index,column=34,value='1')
    wsheet_ll.cell(row=row_index,column=36,value='6')
    #wsheet_ll.cell(row=row_index,column=20,value=cell_Pdis.value)
    if cell_NPS.value is not None:
        nps_value = medidas.get(cell_NPS.value)  
        if nps_value is not None:
            nps_value = int(nps_value) 
            if nps_value <= 25:
                wsheet_ll.cell(row=row_index, column=38, value='Sound Engineering Practice')
                wsheet_ll.cell(row=row_index, column=37, value='4.3')
    #if cell_NPS.value is not None and cell_Pdis.value is not None:
        
        if cell_Pdis.value is not None:  # Check if cell_Pdis.value is not None
            try:
                pdis_value = float(cell_Pdis.value)  # Convert cell_Pdis.value to a float
                if nps_value <= 100 and nps_value * pdis_value <= 1000:
                    wsheet_ll.cell(row=row_index, column=37, value='1')
                    wsheet_ll.cell(row=row_index, column=38, value='A')
                elif 25 < nps_value <= 350 and nps_value * pdis_value <= 3500:
                    wsheet_ll.cell(row=row_index, column=37, value='2')
                    wsheet_ll.cell(row=row_index, column=38, value='A2,D1,E1')
                elif (nps_value > 350 and pdis_value <= 10) or (pdis_value < 10 and nps_value * pdis_value > 3500):
                    wsheet_ll.cell(row=row_index, column=37, value='3')
            except ValueError:
           
                print('ojo')
     

      
line_list_template.save('Resultado_ll.xlsx')