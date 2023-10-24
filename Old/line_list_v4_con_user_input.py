import os, re, openpyxl
from openpyxl.styles  import Alignment
#------------------------------------- PARTE 1 --------------------------------------- CARGAR ARCHIVOS Y ARMAR LA BASE COMPLETANDO LA CLASIFICACION PED PARA LOS DIAMETROS MENORES A 25.-----------------
# ESTE YA FUNCIONA BIEEEEEEEEEEEEEEEEEEEEEN
codigo_proyecto=input('Ingresar codigo del proyecto').upper
nombre_proyecto=input('Ingresar el nombre del proyecto').title 
nombre_planta=input('Ingresar el nombre de la planta').title
# INPUT FILE 
try:
    current_path=os.getcwd()
    input_file_name='Line_list_base.xlsx'
    input_file_path=os.path.join(current_path,input_file_name)
    input_wb=openpyxl.load_workbook(input_file_path)
    input_ws=input_wb.active    

# TEMPLATE FILE 
    template_file_name=f'line_list_template_{nombre_proyecto}.xlsx'
    template_file_path=os.path.join(current_path,template_file_name)
    template_wb=openpyxl.load_workbook(template_file_path)
    template_ws=template_wb['sheet1']    
except FileNotFoundError:
    pass
# Cargo las lineas filtrando con regex. OJO LA EXPRESION DE REGEX ES MUY VAGA. PODRIA COMPLICARCLA.
lineas=[]
medidas={'OD8mm': '8', 'OD12mm': '12', '1/8"': 6, '1/4"': 8, '3/8"': 10, '1/2"': 15, '3/4"': 20, '1"': 25, '1 1/4"': 32, '1 1/2"': 40, '2"': 50, '2 1/2"': 65, '3"': 80, '4"': 100, '6"': 150, '8"': 200, '10"': 250}
patron=r'-\d{4}-'
# Patron solo filtra cadenas que tengan 4 numeros encerrados por guiones. 
for row in input_ws['C']:
    if re.search(patron,row.value):
        lineas.append(row.value)
#print(lineas) Ahoa linea guarda todas las lineas en forma de lista.

row_index=11
for linea in lineas:
    # VOLVER A ESCRIBIR CON PANDAS.
    template_ws.cell(row=row_index, column=4, value=linea.split('-')[0].strip()) # dimension
    template_ws.cell(row=row_index, column=5,value=linea.split('-')[1].strip()) # fluido
    template_ws.cell(row=row_index, column=16,value=linea.split('-')[1].strip()) # fluido
    template_ws.cell(row=row_index, column=6,value=linea.split('-')[2].strip()) # tag
    template_ws.cell(row=row_index, column=7,value=linea.split('-')[3].strip()) # clase cañeria
    if len(linea.split("-")) > 4:
        template_ws.cell(row=row_index, column=8, value=linea.split('-')[4].strip()) # aislamiento
    else:
        template_ws.cell(row=row_index, column=8, value="-")
    if "L" in linea.split('-')[1].strip():
        template_ws.cell(row=row_index, column=17, value='LIQ')

    template_ws.cell(row=row_index,column=35,value='NO')   # Ver si no hay que poner algún condicional para estas líneas. 
    template_ws.cell(row=row_index,column=34,value='1')
    template_ws.cell(row=row_index,column=36,value='6')

    
    row_index+=1
   
# PONGO TITULOS Y FORMATOS.
template_ws.cell(row=4, column=11, value=nombre_proyecto)
template_ws.cell(row=5, column=11, value=nombre_planta)
template_ws.cell(row=6, column=11, value=codigo_proyecto)
template_ws.cell(row=7, column=11, value=f'{codigo_proyecto}-295-01-R{rev} Line List')
template_ws.cell(row=2, column=19, value=f'{codigo_proyecto}-295-01 Line List')
merge_range='S2:AL8'
template_ws.merge_cells(merge_range)
alignment = Alignment(horizontal='center', vertical='center')  
template_ws['S2'].alignment = alignment


row_start=11 # NO SE SI SE PUEDE HACER DE OTRA FORMA PERO ASI FUNCIONA BIEN Y SE EVITAN ERRORES.
template_wb.save('line_list_paso_1.xlsx')
try:    
    parte_2_wb=openpyxl.load_workbook('line_list_paso_2.xlsx')
    parte_2_ws=parte_2_wb.active  
    
    for index in range(row_start, parte_2_ws.max_row+1):
        if parte_2_ws.cell(row=index, column=4).value is not None:
            nps=float(medidas.get(template_ws.cell(row=index, column=4).value))
            Pdis=parte_2_ws.cell(row=index, column=20).value
            #print(f'P {Pdis}  nps {nps}')
            if Pdis==0:
                continue
            if nps<=25:
                parte_2_ws.cell(row=row_index, column=38, value='Sound Engineering Practice')
                parte_2_ws.cell(row=row_index, column=37, value='4.3') 
            elif Pdis * nps <= 1000 and nps <= 100:
                parte_2_ws.cell(row=index, column=37, value='1')
                parte_2_ws.cell(row=index, column=38, value='A')
            elif (nps * Pdis <= 3500) and (25 < nps <= 350):
                parte_2_ws.cell(row=index, column=37, value='2')
                parte_2_ws.cell(row=index, column=38, value='A1,D1,E1')
            elif (nps>350 and Pdis<=10) or (Pdis<10 and Pdis*nps<3500):
                parte_2_ws.cell(row=index, column=37, value='3')
                parte_2_ws.cell(row=index, column=38, value='B1+D,F ; B+E,C1; H')  

            index+=1
    parte_2_wb.save(f'{codigo_proyecto}-295-01-R{rev} Line List.xlsx')
    print(f'paso 2 completo, se genero un archivo con el nombre: {codigo_proyecto}-295-01-R{rev} Line List.xlsx')  
except FileNotFoundError:
    print('Paso 1 completo, se generó el acrhivo line_list_paso_1.xlsx. Completar la columna de presiones de diseño y guardar el archivo como line_list_paso_2.xlsx')
    print('Todavía no se generó el archivo line list_paso_2.xlsx')













    
    


